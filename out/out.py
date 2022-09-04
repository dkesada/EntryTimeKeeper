#! /usr/bin/python
# -*- coding: utf-8 -*-
# author: David Quesada López

"""
Backend. Each time a user arrives, we will show it on an excell file.
It will read a dictionary for the users and codes and then it will create excells
for each day.

Will use Pickle to load a dictionary with the names and codes of the users, 
and OpenPyXL for creating and editing the excels
"""

import pickle
import datetime
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from random import randint

data_name = 'base.p'
file_name = './informes/' + datetime.datetime.now().strftime("%d-%m-%Y") + '.xlsx'


class Backend:
    def __init__(self):
        # a = os.path.dirname(os.path.abspath(__file__))
        with open(data_name, 'rb') as f:
            self.base = pickle.load(f, encoding='latin1')
        self.date = datetime.datetime.today()  # .day , .month , .year

        # Check if the file for today exists
        if not os.path.isfile(file_name):
            self.createFile()

    def createFile(self):
        wb = Workbook()
        ws = wb.active
        for i in self.base.items():
            ws['A' + str(i[1][1])] = i[1][0]  # .decode(encoding='latin1') #badly encoded
        wb.save(file_name)

    def deleteFile(self):
        os.remove(file_name)

    def check(self, code):
        """
		Check if the code exists
		"""
        if code in self.base:
            return True
        else:
            return False

    def entry(self, code):
        """
		First, check if the entry has already been registered.
		If not, register the entry
		"""
        ret = False
        wb = load_workbook(file_name)
        ws = wb['Sheet']

        if ws['D' + str(self.base[code][1])].value is None:
            hora = self.digits(datetime.datetime.now().hour, 10) + ':' + self.digits(datetime.datetime.now().minute, 10)
            ws['D' + str(self.base[code][1])] = hora
            wb.save(file_name)
            ret = True

        return ret

    def exit(self, code):
        """
		First, check if the entry has already been registered.
		If not, register the entry
		"""
        ret = False
        wb = load_workbook(file_name)
        ws = wb['Sheet']

        if ws['E' + str(self.base[code][1])].value is None:
            hora = self.digits(datetime.datetime.now().hour, 10) + ':' + self.digits(datetime.datetime.now().minute, 10)
            ws['E' + str(self.base[code][1])] = hora
            wb.save(file_name)
            ret = True

        return ret

    @staticmethod
    def digits(num, k):
        """ Puts '0' before k digit numbers for format purposes """
        ret = str(num)
        if num < k:
            ret = '0' + str(num)
        return ret

    def delUser(self, code):
        """
		Delete a user from the database
		"""
        usr = self.base.pop(code)  # Remove the user

        for i in self.base.keys():  # Lower the position of the rest of the users above it
            if self.base[i][1] > usr[1]:
                usr_i = self.base[i]
                usr_i = (usr_i[0], usr_i[1] - 1)
                self.base[i] = usr_i

        self.dumpDB()

    def addUser(self, name):
        """
        Add a user to the database
        """
        code = self.digits(randint(0, 99999), 1e4)
        while code in self.base.keys():
            code = self.digits(randint(0, 99999), 1e4)
        self.base[code] = (name, self.base.__len__() + 1)
        self.dumpDB()

        return code

    def dumpDB(self):
        with open(data_name, 'wb') as f:
            pickle.dump(self.base, f)
